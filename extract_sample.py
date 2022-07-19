import os
import json
import copy
import pandas as pd
from openpyxl import Workbook

from typing import List, Dict
from dataclasses import dataclass, field

### Data Structure
@dataclass
class NE_Data:
    text: str = ""
    label: str = ""

@dataclass
class Sample_Data:
    src: str = ""
    id: str = ""
    sent: str = ""
    ne_list: List[NE_Data] = field(default_factory=list)

### Method
#===============================================================
def extract_ne_samples_from_json(target_path: str, target_ne_list: List[str]):
#===============================================================
    ret_sample_list: List[Sample_Data] = []

    if not os.path.exists(target_path):
        print(f"[extract_sample][extract_ne_samples_from_json] ERR - Not Exists: {target_path}")
        return ret_sample_list

    data_source = target_path.split("/")[-1]

    # parse
    with open(target_path, mode="r", encoding="utf-8") as src_file:
        src_json = json.load(src_file)

        doc_arr = src_json["document"]
        for doc_obj in doc_arr:
            sent_arr = doc_obj["sentence"]
            for sent_obj in sent_arr:
                sample_data = Sample_Data(src=data_source,
                                          id=sent_obj["id"],
                                          sent=sent_obj["form"])

                # NE
                is_need_add = False
                ne_arr = sent_obj["NE"]
                for ne_obj in ne_arr:
                    ne_obj_txt = ne_obj["form"]
                    ne_obj_type = ne_obj["label"]
                    ne_data = NE_Data(text=ne_obj_txt,
                                      label=ne_obj_type)
                    sample_data.ne_list.append(copy.deepcopy(ne_data))

                    for target_ne in target_ne_list:
                        if "*" in target_ne:
                            lhs_target_ne = target_ne.split("_")[0]
                            if lhs_target_ne in ne_obj_type:
                                is_need_add = True
                        else:
                            if target_ne == ne_obj_type:
                                is_need_add = True
                if is_need_add:
                    ret_sample_list.append(sample_data)

    return ret_sample_list

#===============================================================
def make_target_ne_dict_extract_results(data_list: List[Sample_Data], target_ne_list: List[str]):
#===============================================================
    # init
    ne_data_dict = {target_ne: [] for target_ne in target_ne_list}

    # arrange
    for sample_data in data_list:
        is_insert = False
        for ne_data in sample_data.ne_list:
            if is_insert:
                break
            for target_ne in target_ne_list:
                if "*" in target_ne:
                    lhs_target_ne = target_ne.split("_")[0]
                    if lhs_target_ne in ne_data.label:
                        ne_data_dict[target_ne].append(sample_data)
                        is_insert = True
                        break
                else:
                    if target_ne == ne_data.label:
                        ne_data_dict[target_ne].append(sample_data)
                        is_insert = True
                        break

    # print (key, values) size
    dict_values_size = 0
    for target_ne in target_ne_list:
        print(f"[extract_sample][make_excel_file_using_extract_results] {target_ne}.size: {len(ne_data_dict[target_ne])}")
        dict_values_size += len(ne_data_dict[target_ne])
    print(f"[extract_sample][make_excel_file_using_extract_results] total dict.values.size {dict_values_size}")

    return ne_data_dict

#===============================================================
def make_excel_file_using_extract_results(target_ne_dict: Dict[str, List[Sample_Data]]):
#===============================================================
    # create sheet
    new_wb = Workbook()
    work_sheet = new_wb.active

    for dic_idx, (key, sample_list) in enumerate(target_ne_dict.items()):
        print(f"[extract_sample][make_excel_file_using_extract_results] {key} is Processing... (size: {len(sample_list)})")
        if key != work_sheet.title:
            if "*" in key:
                work_sheet.title = key.replace("*", "all")
            else:
                work_sheet.title = key

        row = 2
        col = 2
        work_sheet.cell(row, col, value="domain")
        work_sheet.cell(row, col+1, value="id")
        work_sheet.cell(row, col+2, value="sentence")
        work_sheet.cell(row, col+3, value="ne_text")
        work_sheet.cell(row, col+4, value="ne_label")

        row += 1
        for sp_idx, sample_data in enumerate(sample_list):
            work_sheet.cell(row, col, value=sample_data.src)
            work_sheet.cell(row, col+1, value=sample_data.id)
            work_sheet.cell(row, col+2, value=sample_data.sent)

            for ne_item in sample_data.ne_list:
                work_sheet.cell(row, col+3, ne_item.text)
                work_sheet.cell(row, col+4, ne_item.label)
                row += 1
            row += 1

        # new sheet
        if dic_idx != len(target_ne_dict.keys())-1:
            work_sheet = new_wb.create_sheet()

    # save *.xlsx
    save_path = "./target_samples.xlsx"
    new_wb.save(save_path)
    print(f"[extract_sample][make_excel_file_using_extract_results] Complete - {save_path}")

### Main
if "__main__" == __name__:
    print(f"[extract_sample][__main__] ----MAIN")
    # init
    '''
        @NOTE:
            '*'이 붙어 있으면 모든 세분류에 대해 검색.
    '''
    target_ne_list: List[str] = ["PS_NAME",
                                 "OG_*",
                                 "LC_*",
                                 "QT_*",
                                 "CV_OCCUPATION"]
    print(f"[extract_sample][__main__] {target_ne_list}")

    # load
    target_dir_path = "./data/old_nikl"
    target_json_list = os.listdir(target_dir_path)
    target_json_list = [target_dir_path+"/"+x for x in target_json_list if ".json" in x]
    print(f"[extract_sample][__main__] target json list: {target_json_list}")

    # extract
    all_sample_data_list: List[Sample_Data] = []
    for target_path in target_json_list:
        print(f"[extract_sample][__main__] curr target: {target_path}")

        extract_data_list = extract_ne_samples_from_json(target_path=target_path, target_ne_list=target_ne_list)
        print(f"[extract_sample][__main__] extract data size: {len(extract_data_list)}")

        all_sample_data_list.extend(extract_data_list)
    print(f"[extract_sample][__main__] all sample data size: {len(all_sample_data_list)}")

    # make target_ne_dict
    target_ne_dict = make_target_ne_dict_extract_results(data_list=all_sample_data_list,
                                                         target_ne_list=target_ne_list)

    # make excel
    make_excel_file_using_extract_results(target_ne_dict)